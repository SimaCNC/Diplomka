import openpyxl
from openpyxl import workbook.
from openpyxl.styles import PatternFill, Border, Side, Alignment


#nahrani template do pameti
cesta_template = "Datasheet_senzor_template.xlsx"
template = openpyxl.load_workbook(cesta_template)

#nazev souboru a jeho vytvoreni
filename = "data"
template.save(filename)

wb = openpyxl.load_workbook(filename)
ws = wb["MD0001"]

